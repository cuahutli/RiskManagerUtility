from openpyxl import load_workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

wb = load_workbook('basefiles/credito_actual.xlsx')
wsc = wb['Hoja1']
wsp = wb.create_sheet('productos')

column = ''
headers = {get_column_letter(cell.column): cell.value for cell in wsc[1]}
for key, value in headers.items():
    if value == 'PRODUCTO DE CREDITO':
        column = key
# print(headers)
sucursales = [data.value for data in wsc[column] if data.value != 'PRODUCTO DE CREDITO']
print(list(dict.fromkeys(sucursales)))
